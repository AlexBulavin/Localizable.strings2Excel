# -*- coding:utf-8 -*-
# First install library
# pip install openpyxl

from optparse import OptionParser
from XlsFileUtil import XlsFileUtil
from XmlFileUtil import XmlFileUtil
from StringsFileUtil import StringsFileUtil
from Log import Log
import os
import time
import openpyxl


def addParser():
    parser = OptionParser()

    parser.add_option("-f", "--fileDir",
                      help="Xls files directory.",
                      metavar="fileDir")

    parser.add_option("-t", "--targetDir",
                      help="The directory where the strings files will be saved.",
                      metavar="targetDir")

    parser.add_option("-e", "--excelStorageForm",
                      type="string",
                      default="multiple",
                      help="The excel(.xls) file storage forms including single(single file), multiple(multiple files), default is multiple.",
                      metavar="excelStorageForm")

    parser.add_option("-a", "--additional",
                      help="additional info.",
                      metavar="additional")

    (options, args) = parser.parse_args()
    Log.info("options: %s, args: %s" % (options, args))

    return options


def convertFromSingleForm(options, fileDir, targetDir):

    for filenames in os.listdir(fileDir):

        wb = openpyxl.load_workbook(fileDir + "/" + filenames)
        ws = wb.active
        cols = ws.max_column
        rows = ws.max_row
        keys = []

        for cell_in_row in range(1, cols+1):  # Pass through row one by one column
            values = []

            for rows_in_column in range(2, rows + 1):  # Pass through the rows
                if cell_in_row == 1:  # Check is the row index = 1. If yes, then append key
                    val = ws.cell(row=rows_in_column, column=1).value
                    if val is not None:
                        keys.append(val)
                else:
                    val = ws.cell(row=rows_in_column, column=cell_in_row).value
                    if val is not None:
                        values.append(val)

            if cell_in_row > 1:
                languageName = ws.cell(row=1, column=cell_in_row).value
                if not languageName.endswith(".lproj"): # Use this line of code if first does not have .lproj extension in column names
                    dir_and_file = languageName + ".lproj/"
                else:
                    dir_and_file = languageName + "/"  # Use this line of code if first row already have .lproj extension
                StringsFileUtil.writeToFile(
                    keys,
                    values,
                    targetDir + "/" + dir_and_file,
                    filenames.replace(".xls",
                                      "") + ".strings",
                    options.additional)

    print("%s" %cols, "files converted to %s successfully!" %fileDir)
    print("You can see strings files in %s" %targetDir)


def convertFromMultipleForm(options, fileDir, targetDir):
    for filenames in os.listdir(fileDir):
        #TODO For Python 3 this method need to be refactor accordingly to convertFromSingleForm
        xlsFilenames = [fi for fi in filenames if fi.endswith(".xls")]
        for file in xlsFilenames:
            xlsFileUtil = XlsFileUtil(fileDir + "/" + file)
            langFolderPath = targetDir + "/" + file.replace(".xls", "")
            if not os.path.exists(langFolderPath):
                os.makedirs(langFolderPath)

            for sheet in xlsFileUtil.getAllTables():
                iosDestFilePath = langFolderPath + "/" + sheet.name
                iosFileManager = open(iosDestFilePath, "wb")
                for row in sheet.get_rows():
                    content = "\"" + row[0].value + "\" " + \
                              "= " + "\"" + row[1].value + "\";\n"
                    iosFileManager.write(content)
                if options.additional is not None:
                    iosFileManager.write(options.additional)
                iosFileManager.close()
    print("Convert %s successfully! you can see strings file in %s" % (
        fileDir, targetDir))


def startConvert(options):
    fileDir = options.fileDir
    targetDir = options.targetDir

    print("Start converting")

    if fileDir is None:
        print("xls files directory can not be empty! try -h for help.")
        return

    if targetDir is None:
        print("Target file directory can not be empty! try -h for help.")
        return

    targetDir = targetDir + "/xls-files-to-strings_" + \
                time.strftime("%Y%m%d_%H%M%S")
    if not os.path.exists(targetDir):
        os.makedirs(targetDir)
    if options.excelStorageForm == "single":
        convertFromSingleForm(options, fileDir, targetDir)
    else:
        convertFromMultipleForm(options, fileDir, targetDir)


def main():
    options = addParser()
    startConvert(options)


if __name__ == '__main__':
    main()
