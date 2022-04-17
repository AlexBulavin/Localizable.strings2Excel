__author__ = 'Alex Bulavin'
# Backup for Xml2Xls.py file

# -*- coding:utf-8 -*-

import os
from optparse import OptionParser
from XmlFileUtil import XmlFileUtil
import pyExcelerator
from Log import Log
import time
import openpyxl


def addParser():
    parser = OptionParser()

    parser.add_option("-f", "--fileDir",
                      help="strings.xml files directory.",
                      metavar="file_dir")

    parser.add_option("-t", "--targetDir",
                      help="The directory where the xls files will be saved.",
                      metavar="target_dir")

    parser.add_option("-e", "--excelStorageForm",
                      type="string",
                      default="multiple",
                      help="The excel(.xls) file storage forms including single(single file), multiple(multiple files), default is multiple.",
                      metavar="excelStorageForm")

    (options, args) = parser.parse_args()
    Log.info("options: %s, args: %s" % (options, args))

    return options


def convertToMultipleFiles(file_dir, target_dir):
    dest_dir = genDestDir(target_dir)

    for dirnames in os.walk(file_dir):
        valuesDirs = [di for di in dirnames if di.startswith("values")]
        for dirname in valuesDirs:
            workbook = pyExcelerator.Workbook()
            for filenames in os.walk(file_dir + '/' + dirname):
                xmlFiles = [fi for fi in filenames if fi.endswith(".xml")]
                for xmlfile in xmlFiles:
                    ws = workbook.add_sheet(xmlfile)
                    path = file_dir + '/' + dirname + '/' + xmlfile
                    (keys, values) = XmlFileUtil.getKeysAndValues(path)
                    for keyIndex in range(len(keys)):
                        key = keys[keyIndex]
                        value = values[keyIndex]
                        ws.write(keyIndex, 0, key)
                        ws.write(keyIndex, 1, value)
            file_path = dest_dir + "/" + getCountryCode(dirname) + ".xls"
            workbook.save(file_path)
    print("Convert %s successfully! you can see xls file in %s" % (
        file_dir, dest_dir))


def convertToSingleFile(file_dir, target_dir):
    values_dirs = []
    xml_files = []
    dest_dir = genDestDir(target_dir)
    # for dirnames in os.listdir(file_dir + '/'):
    #     print("dirnames = ", dirnames)
    #     values_dirs = [di for di in dirnames if di.startswith("values")]
    for dirname in os.listdir(file_dir + '/'):  # Loop through upper level of parsing directory
        dirnames = os.listdir(file_dir)
        # print("dirnames = ", dirnames)
        values_dirs.append(dirname)
        # print("dirname = ", dirname)
        for filenames in os.listdir(file_dir + '/' + dirname):  # Loop through inner level of each parsing directory
            # print("filenames = ", filenames)
            # xml_files += [fi for fi in filenames if fi.endswith(".xml")]
            # print("xml_files", xml_files)
            # for xml_file in xml_files:
            file_name = filenames.replace(".xml", "")
            # print("file_name = ", file_name)
            file_path = dest_dir + "/" + file_name + ".xls"
            output_file_path = dest_dir + "/" + dirname + "/" + file_name + ".xls"
            os.makedirs(dest_dir + "/" + dirname)
            # print("file_path = ", file_path)
            # print("output_file_path", output_file_path)
            if not os.path.exists(output_file_path):
                workbook = openpyxl.Workbook()
                ws = workbook.create_sheet(title="Xml 2 xls strings", index=0)
                index = 1
                # print("ws = ", ws)
                ws.cell(row=1, column=1).value = 'keyName'
                # print("ws.cell(row=1, column=1).value = ", ws.cell(row=1, column=1).value)
                for dirname1 in values_dirs:
                    ws.cell(row=1, column=index+1).value = getCountryCode(dirname1)
                    path = file_dir + dirname1 + '/' + filenames
                    # print("path for XmlFileUtil.getKeysAndValues(path) = ", path)
                    (keys, values) = XmlFileUtil.getKeysAndValues(path)
                    print("keys = ", keys)
                    for x in range(len(keys)):
                        key = keys[x]
                        value = values[x]
                        if index == 1:
                            ws.cell(row=x+2, column=1).value = key
                            ws.cell(row=x+2, column=2).value = value
                        else:
                            ws.cell(row=x+2, column=index + 1).value = value
                    index += 1
                workbook.save(output_file_path)
    print("Convert %s successfully! you can see xls file in %s" % (
        file_dir, dest_dir))


def genDestDir(target_dir):
    dest_dir = target_dir + "/xml-files-to-xls_" + \
              time.strftime("%Y%m%d_%H%M%S")
    if not os.path.exists(dest_dir):
        os.makedirs(dest_dir)
    return dest_dir


def getCountryCode(dirname):
    code = 'en'
    dir_split = dirname.split('values-')
    if len(dir_split) > 1:
        code = dir_split[1]
    return code


def startConvert(options):
    file_dir = options.fileDir
    target_dir = options.targetDir

    print("Start converting")

    if file_dir is None:
        print("strings.xml files directory can not be empty! try -h for help.")
        return

    if target_dir is None:
        print("Target file path can not be empty! try -h for help.")
        return

    if options.excelStorageForm == "single":
        convertToSingleFile(file_dir, target_dir)
    else:
        convertToMultipleFiles(file_dir, target_dir)


def main():
    options = addParser()
    startConvert(options)


if __name__ == '__main__':
    main()
