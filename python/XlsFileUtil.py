#!/usr/bin/python
# -*- coding: UTF-8 -*-


import sys
from importlib import reload
from openpyxl import load_workbook

import os


class XlsFileUtil:
    """xls file util"""

    def __init__(self, filePath):
        self.filePath = filePath

        reload(sys)
        self.data = load_workbook(filename=filePath)
        self.sheet = self.data.active
        self.rows_count = self.sheet.max_row
        self.columns_count = self.sheet.max_column

    def getAllTables(self):
        return self.data.sheetnames

    def getTableByIndex(self, index):
        if 0 <= index < len(self.data.sheetnames):
            return self.data.sheetnames[index]
        else:
            print("XlsFileUtil error -- getTable:index")


    def getTableByName(self, name):
        return self.data.sheetnames(name)  # get_sheet_names(name)
